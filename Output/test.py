from openpyxl import load_workbook

# 1. reading an existing workbook
wb= load_workbook('./material/practice1.xlsx')
# 2. get the table
staff_sheet = wb['下半年公司名单']

# 3. get and update the specified data
for cell in staff_sheet.iter_rows(min_col=4, max_col=4, values_only=True):
    if(cell.value == '部门'):
        continue
    print(cell.value)
    cell.value = '战略储备部'

# 4. save the sheet
staff_sheet.save('practice2_result.xlsx')